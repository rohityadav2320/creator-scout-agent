import asyncio
import csv
import json
import os
import random
import re
import time
from datetime import datetime
from playwright.async_api import async_playwright

RESULTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "results")
MASTER_CSV = os.path.join(RESULTS_DIR, "creators_master.csv")
MASTER_XLSX = os.path.join(RESULTS_DIR, "creators_master.xlsx")
MASTER_COLS = ["reel_url", "username", "full_name", "likes", "comments",
               "hashtags", "caption", "first_scraped", "last_seen"]


def _update_master(reels):
    """Accumulate every reel ever scraped into a single de-duplicated master
    CSV + Excel (keyed by reel_url). New reels are appended; reels seen again
    have their like/comment counts refreshed. This is the permanent local store."""
    os.makedirs(RESULTS_DIR, exist_ok=True)
    existing = {}
    if os.path.exists(MASTER_CSV):
        try:
            with open(MASTER_CSV, newline="", encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    existing[row.get("reel_url", "")] = row
        except Exception as e:
            debug_log(f"_update_master: could not read existing master: {e}")

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    added = 0
    for r in reels:
        url = r.get("reel_url", "")
        if not url:
            continue
        if url in existing:
            existing[url]["likes"] = r.get("likes", 0)
            existing[url]["comments"] = r.get("comments", 0)
            existing[url]["last_seen"] = now
        else:
            existing[url] = {
                "reel_url": url,
                "username": r.get("username", ""),
                "full_name": r.get("full_name", ""),
                "likes": r.get("likes", 0),
                "comments": r.get("comments", 0),
                "hashtags": ", ".join(r.get("hashtags", []) or []),
                "caption": (r.get("caption", "") or "").replace("\n", " ")[:500],
                "first_scraped": now,
                "last_seen": now,
            }
            added += 1

    rows = list(existing.values())

    # Master CSV
    with open(MASTER_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=MASTER_COLS)
        w.writeheader()
        for row in rows:
            w.writerow({c: row.get(c, "") for c in MASTER_COLS})

    # Master Excel (clickable reel links)
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Creators"
        ws.append(MASTER_COLS)
        for row in rows:
            ws.append([row.get(c, "") for c in MASTER_COLS])
            cell = ws.cell(row=ws.max_row, column=1)
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
        wb.save(MASTER_XLSX)
    except Exception as e:
        debug_log(f"_update_master: xlsx write failed: {e}")

    debug_log(f"_update_master: +{added} new, {len(rows)} total in master.")
    return added, len(rows)


def save_reels_csv(reels, scraped_by="", batch_id="", scraped_from="", push_db=True):
    """Write EVERY scraped reel to disk immediately — independent of the UI or any
    filtering. Always produces results/all_reels_latest.csv (overwritten each run),
    a timestamped copy, and results/links_latest.txt (just the URLs). Pushes to the
    shared Supabase bank ONLY if push_db=True — triage pushes the kept ones later."""
    if not reels:
        debug_log("save_reels_csv: no reels to save.")
        return None
    os.makedirs(RESULTS_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    cols = ["reel_url", "username", "full_name", "likes", "comments", "views",
            "followers", "category", "contact_email", "bio", "external_url",
            "hashtags", "caption", "audio", "scraped_by", "scraped_from", "batch_id"]
    rows = []
    for r in reels:
        row = dict(r)
        row["hashtags"] = ", ".join(r.get("hashtags", []) or [])
        row["caption"] = (r.get("caption", "") or "").replace("\n", " ")[:500]
        row["scraped_by"] = scraped_by
        row["scraped_from"] = scraped_from
        row["batch_id"] = batch_id
        rows.append({c: row.get(c, "") for c in cols})

    for path in (os.path.join(RESULTS_DIR, "all_reels_latest.csv"),
                 os.path.join(RESULTS_DIR, f"all_reels_{stamp}.csv")):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            w.writerows(rows)

    with open(os.path.join(RESULTS_DIR, "links_latest.txt"), "w", encoding="utf-8") as f:
        for r in reels:
            f.write(f"{r.get('reel_url','')}\t@{r.get('username','')}\t{r.get('likes',0)} likes\n")

    # Update the permanent, accumulating master CSV + Excel.
    added, total = _update_master(reels)

    # Push to Supabase only when asked (triage saves the KEPT creators later).
    if push_db:
        try:
            import db
            if db.is_configured():
                new_n, known_n, err = db.upsert_reels(
                    reels, scraped_by=scraped_by, batch_id=batch_id, scraped_from=scraped_from)
                if err:
                    debug_log(f"Supabase upsert error: {err}")
                else:
                    debug_log(f"Supabase: {new_n} new, {known_n} already known.")
        except Exception as e:
            debug_log(f"Supabase push skipped: {e}")

    debug_log(f"save_reels_csv: wrote {len(reels)} reels; master +{added} new ({total} total)")
    return os.path.join(RESULTS_DIR, "all_reels_latest.csv")

INSTAGRAM_URL = "https://www.instagram.com"
DEBUG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "debug")


def debug_log(msg):
    os.makedirs(DEBUG_DIR, exist_ok=True)
    ts = datetime.now().strftime("%H:%M:%S")
    with open(os.path.join(DEBUG_DIR, "run.log"), "a") as f:
        f.write(f"[{ts}] {msg}\n")


def debug_dump_response(url, body, idx):
    os.makedirs(DEBUG_DIR, exist_ok=True)
    safe = re.sub(r"[^a-zA-Z0-9]+", "_", url.split("?")[0])[-60:]
    path = os.path.join(DEBUG_DIR, f"resp_{idx:03d}_{safe}.json")
    try:
        with open(path, "w") as f:
            json.dump(body, f, indent=2)
        debug_log(f"Dumped response from {url[:120]}")
    except Exception as e:
        debug_log(f"Failed to dump response from {url[:120]}: {e}")


# Each Instagram account gets its own persistent browser profile so the feed
# stays trained across scrape runs.  All browser state (cookies, localStorage,
# history, fingerprint) is written to disk automatically by Playwright.
_DATA_DIR = os.environ.get("CREATOR_SCOUT_DATA_DIR") or os.path.dirname(os.path.abspath(__file__))
PROFILES_DIR = os.path.join(_DATA_DIR, "browser_profiles")

# Real Chrome binary — passes Instagram's bot detection far better than
# Playwright's bundled Chromium which exposes automation fingerprints.
_CHROME_PATHS = [
    "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",  # macOS
    "/usr/bin/google-chrome",                                          # Linux
    "/usr/bin/chromium-browser",                                       # Linux fallback
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",         # Windows
]
_CHROME_EXECUTABLE = next((p for p in _CHROME_PATHS if os.path.exists(p)), None)


def _get_profile_dir(username):
    """Return the on-disk profile folder for this Instagram account."""
    safe = re.sub(r"[^\w.-]", "_", username.lower()) if username else "default"
    path = os.path.join(PROFILES_DIR, safe)
    os.makedirs(path, exist_ok=True)
    return path


async def _apply_stealth(page):
    """Patch the page to hide Playwright automation signals from Instagram."""
    try:
        from playwright_stealth import Stealth
        await Stealth().apply_stealth_async(page)
    except Exception:
        # playwright-stealth not installed — apply minimal manual patches.
        await page.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            window.chrome = { runtime: {} };
            Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3]});
            Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
        """)


_CDP_URL = "http://localhost:9222"
_cdp_available = None  # cached after first check


async def _check_cdp():
    """Return True if a real Chrome is listening on port 9222."""
    import aiohttp
    try:
        async with aiohttp.ClientSession() as s:
            async with s.get(f"{_CDP_URL}/json/version", timeout=aiohttp.ClientTimeout(total=2)) as r:
                return r.status == 200
    except Exception:
        return False


async def _launch_context(p, username=""):
    """Launch Playwright Chromium with a persistent profile (simple & reliable).
    The browser window opens; if not logged in, the user logs in by hand."""
    profile_dir = _get_profile_dir(username)
    debug_log("Using Playwright Chromium with persistent profile")
    context = await p.chromium.launch_persistent_context(
        profile_dir,
        headless=False,
        slow_mo=50,
        viewport={"width": 1280, "height": 900},
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-first-run",
            "--no-default-browser-check",
        ],
    )
    async def _stealth_new_page(page):
        await _apply_stealth(page)
    context.on("page", _stealth_new_page)
    page = context.pages[0] if context.pages else await context.new_page()
    await _apply_stealth(page)
    return context, page


async def _has_session(context):
    """Instagram sets a non-empty `sessionid` cookie only after a successful login."""
    try:
        cookies = await context.cookies()
        return any(c.get("name") == "sessionid" and c.get("value") for c in cookies)
    except Exception:
        return False


async def _session_works(page):
    """Actually verify the session is authenticated for API access — not just that a
    sessionid cookie exists. A stale/checkpointed/expired session redirects API calls
    to the home page and returns HTML. Returns True only on a real JSON response."""
    try:
        if "instagram.com" not in (page.url or ""):
            await safe_goto(page, INSTAGRAM_URL + "/")
            await page.wait_for_timeout(1500)
        result = await page.evaluate(
            """async () => {
                try {
                    const r = await fetch('/api/v1/users/web_profile_info/?username=instagram', {
                        headers: { 'x-ig-app-id': '936619743392459', 'accept': '*/*' },
                        credentials: 'include'
                    });
                    const body = await r.text();
                    let authed = false;
                    try { authed = !!JSON.parse(body)?.data?.user?.id; } catch (e) {}
                    return { status: r.status, redirected: r.redirected,
                             head: body.slice(0, 1), authed: authed };
                } catch (e) { return { status: -1, redirected: false, head: 'E', authed: false }; }
            }"""
        )
        # Truly authenticated only if we got 200 JSON with real user data. A
        # half-injected/dead session returns a 4xx "login_required" JSON or an
        # HTML redirect to the home page.
        ok = (bool(result) and result.get("status") == 200
              and result.get("authed") is True)
        debug_log(f"session check: {result} → works={ok}")
        return ok
    except Exception as e:
        debug_log(f"session check error: {e}")
        return False


async def dismiss_popups(page):
    """Dismiss 'Save login info' and 'Turn on notifications' dialogs if present."""
    for _ in range(2):
        try:
            not_now = page.get_by_role("button", name=re.compile("not now", re.IGNORECASE))
            if await not_now.count() > 0 and await not_now.first.is_visible():
                await not_now.first.click()
                await page.wait_for_timeout(1200)
        except Exception:
            pass


async def safe_goto(page, url, timeout=60000):
    """Navigate without crashing on slow loads. Instagram is heavy and can exceed
    the default 30s; we wait longer and, if it still doesn't fire, carry on — the
    page is usually interactive enough (and cookies are readable) regardless."""
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=timeout)
        return True
    except Exception as e:
        debug_log(f"safe_goto: {url} slow/incomplete ({type(e).__name__}); continuing.")
        return False


async def inject_session_cookie(context, session_id: str):
    """Inject the user's Instagram cookies into the browser context.

    Accepts EITHER:
      • a full cookie string  "ig_did=..; mid=..; csrftoken=..; ds_user_id=..; sessionid=.."
        (best — replicates the real browser's device identity so Instagram accepts it)
      • a bare sessionid       (we derive ds_user_id from its first segment)

    A bare sessionid relies on the persistent profile's existing device cookies
    (mid/ig_did/datr) for Instagram to trust the session — so we DON'T clear them.
    add_cookies overwrites the sessionid for the same domain/path; everything else
    (the device identity the profile built up) is preserved. Clearing them was a
    regression that made IG distrust the session and bounce to the login page."""
    import urllib.parse
    raw = session_id.strip()

    cookies = []
    is_full = ("sessionid=" in raw) or (raw.count("=") >= 2)
    if is_full:
        # Parse a full "name=value; name=value" cookie header.
        for part in raw.split(";"):
            part = part.strip()
            if "=" not in part:
                continue
            name, _, value = part.partition("=")
            name, value = name.strip(), value.strip()
            if name and value:
                cookies.append({
                    "name": name, "value": value, "domain": ".instagram.com",
                    "path": "/", "secure": True, "sameSite": "Lax",
                })
        names = [c["name"] for c in cookies]
        debug_log(f"Injected FULL cookie set: {names}")
    else:
        # Bare sessionid → also inject the matching ds_user_id (first segment).
        decoded = urllib.parse.unquote(raw)
        ds_user_id = decoded.split(":")[0] if ":" in decoded else ""
        cookies.append({
            "name": "sessionid", "value": raw, "domain": ".instagram.com",
            "path": "/", "httpOnly": True, "secure": True, "sameSite": "Lax",
        })
        if ds_user_id.isdigit():
            cookies.append({
                "name": "ds_user_id", "value": ds_user_id, "domain": ".instagram.com",
                "path": "/", "secure": True, "sameSite": "Lax",
            })
        debug_log(f"Injected BARE sessionid (ds_user_id={ds_user_id or 'UNKNOWN'}) "
                  f"— may fail device check; full cookie string preferred.")

    if cookies:
        await context.add_cookies(cookies)


async def ensure_logged_in(page, context, username=None, progress_callback=None,
                           timeout_seconds=600, session_id=None):
    """
    Preferred: if session_id is provided, inject it directly (no browser login needed).
    Fallback: check saved profile cookies, then open the login page for manual login.
    """
    # ── Fast path: cookie injection (SIMPLE — trust the cookie, like before) ──
    if session_id:
        await inject_session_cookie(context, session_id)
        await safe_goto(page, f"{INSTAGRAM_URL}/")
        await page.wait_for_timeout(2000)
        if await _has_session(context):
            debug_log("Cookie injection successful (sessionid present).")
            await dismiss_popups(page)
            return True
        debug_log("Injected cookie not present — falling through to manual login.")

    # ── Saved profile session ─────────────────────────────────────────────────
    await safe_goto(page, f"{INSTAGRAM_URL}/")
    await page.wait_for_timeout(2500)
    if await _has_session(context):
        debug_log("Existing session present — skipping login.")
        await dismiss_popups(page)
        return True

    # ── Manual login fallback ─────────────────────────────────────────────────
    await safe_goto(page, f"{INSTAGRAM_URL}/accounts/login/")
    await page.wait_for_timeout(1500)
    if username:
        try:
            await page.wait_for_selector('input[name="username"]', timeout=8000)
            await page.fill('input[name="username"]', username)
        except Exception:
            pass

    if progress_callback:
        progress_callback(
            "⏳ Log in manually in the browser window (type your password, solve any "
            "verification). Scraping starts automatically once you're logged in..."
        )
    debug_log("Waiting for manual login (polling for sessionid cookie)...")

    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        if await _has_session(context):
            debug_log("Manual login detected.")
            await page.wait_for_timeout(2000)
            await dismiss_popups(page)
            return True
        await page.wait_for_timeout(2000)

    debug_log("Manual login timed out.")
    return False


def _looks_like_media(d):
    """A dict that looks like an Instagram media/reel object."""
    if not isinstance(d, dict):
        return False
    has_code = "code" in d or "shortcode" in d or "pk" in d
    has_owner = "user" in d or "owner" in d
    has_stats = any(k in d for k in ("like_count", "comment_count", "play_count",
                                     "view_count", "caption", "media_type"))
    return has_code and has_owner and has_stats


def _deep_find_media(obj, found, depth=0, seen=None):
    """Recursively collect media-like dicts from ANY response shape. This makes the
    parser resilient across home feed / hashtag / explore / profile payloads without
    hard-coding Instagram's ever-changing key names."""
    if depth > 14 or len(found) > 2000:
        return
    if isinstance(obj, dict):
        if _looks_like_media(obj):
            found.append(obj)
            # Still descend into a carousel's children, but not arbitrary nesting.
            for k in ("carousel_media",):
                if isinstance(obj.get(k), list):
                    for v in obj[k]:
                        _deep_find_media(v, found, depth + 1, seen)
            return
        for v in obj.values():
            _deep_find_media(v, found, depth + 1, seen)
    elif isinstance(obj, list):
        for v in obj:
            _deep_find_media(v, found, depth + 1, seen)


async def extract_reel_data_from_response(response_data):
    reels = []
    try:
        items = []
        if isinstance(response_data, dict):
            if "items" in response_data and isinstance(response_data["items"], list):
                items = response_data["items"]
            elif "sections" in response_data and isinstance(response_data["sections"], list):
                # Classic hashtag / explore feed shape: sections[].layout_content.medias[].media
                for sec in response_data["sections"]:
                    lc = (sec or {}).get("layout_content", {}) or {}
                    for m in lc.get("medias", []) or []:
                        if isinstance(m, dict) and m.get("media"):
                            items.append(m["media"])
            elif "data" in response_data and isinstance(response_data["data"], dict):
                data = response_data["data"]
                # Instagram renames this key over time, e.g.
                #   xdt_api__v1__clips__home__connection_v2 (home feed, Jun 2026)
                # Match clips/reels keys only — NOT timeline/feed keys like
                # xdt_api__v1__feed__timeline__connection which returns home-feed posts.
                connection = None
                for key, val in data.items():
                    if (isinstance(val, dict) and "edges" in val
                            and ("clips" in key or "reels" in key)):
                        connection = val
                        break
                if connection:
                    items = [e.get("node", {}).get("media", {})
                             for e in connection.get("edges", []) if isinstance(e, dict)]
                elif isinstance(data.get("xdt_api__v1__media__shortcode__web_info"), dict):
                    items = data["xdt_api__v1__media__shortcode__web_info"].get("items", [])

        # Fallback: if the structured paths found nothing, deep-search the payload.
        if not items:
            _deep_find_media(response_data, items)

        for item in items:
            if not item:
                continue
            # Some shapes wrap the media one level deeper.
            if isinstance(item, dict) and "media" in item and _looks_like_media(item["media"]):
                item = item["media"]
            reel = parse_reel_item(item)
            if reel:
                reels.append(reel)
    except Exception as e:
        debug_log(f"extract_reel_data_from_response error: {e}")
    return reels


def _media_to_creator_record(item):
    """Turn ANY media object (reel OR photo) into a creator record. Used for
    hashtag/explore results where we want the creator regardless of post type."""
    try:
        if not isinstance(item, dict):
            return None
        code = item.get("code") or item.get("shortcode", "")
        user = item.get("user") or item.get("owner") or {}
        username = user.get("username", "")
        if not code or not username:
            return None
        media_type = item.get("media_type", 0)
        path = "reel" if media_type == 2 else "p"
        caption_obj = item.get("caption") or {}
        caption_text = caption_obj.get("text", "") if isinstance(caption_obj, dict) else str(caption_obj or "")
        return {
            "reel_url": f"{INSTAGRAM_URL}/{path}/{code}/",
            "username": username,
            "full_name": user.get("full_name", ""),
            "followers": user.get("follower_count", 0) or 0,
            "profile_pic": user.get("profile_pic_url", ""),
            "caption": caption_text,
            "hashtags": re.findall(r"#(\w+)", caption_text),
            "views": item.get("view_count") or item.get("play_count") or 0,
            "likes": item.get("like_count", 0) or 0,
            "comments": item.get("comment_count", 0) or 0,
            "audio": "",
            "audio_id": "",
            "posted_at": item.get("taken_at", 0),
            "media_type": media_type,
        }
    except Exception:
        return None


def extract_explore_media(data):
    """Extract creator records from the explore/hashtag GraphQL connection. Instagram
    serves hashtag content under *_feed__timeline__connection with each post wrapped
    in node.explore_story.media (or node.media). This is SEPARATE from the trained-feed
    parser, which intentionally skips timeline keys."""
    out = []
    try:
        if not isinstance(data, dict):
            return out
        d = data.get("data") if isinstance(data.get("data"), dict) else data
        for key, val in (d or {}).items():
            if not (isinstance(val, dict) and isinstance(val.get("edges"), list)):
                continue
            if "connection" not in key:
                continue
            for edge in val["edges"]:
                node = (edge or {}).get("node", {}) or {}
                media = node.get("media")
                if not media and isinstance(node.get("explore_story"), dict):
                    media = node["explore_story"].get("media")
                rec = _media_to_creator_record(media) if media else None
                if rec:
                    out.append(rec)
    except Exception as e:
        debug_log(f"extract_explore_media error: {e}")
    return out


def parse_reel_item(item):
    try:
        code = item.get("code") or item.get("shortcode", "")
        if not code:
            return None
        # media_type: 1 = photo, 2 = video/reel, 8 = carousel. Only keep videos.
        media_type = item.get("media_type", 2)
        if media_type not in (2, 0):   # 0 = unknown shape, let it through
            return None

        reel_url = f"{INSTAGRAM_URL}/reel/{code}/"

        user = item.get("user") or item.get("owner", {})
        username = user.get("username", "")
        full_name = user.get("full_name", "")
        followers = user.get("follower_count", 0)
        profile_pic = user.get("profile_pic_url", "")

        caption_obj = item.get("caption") or {}
        caption_text = ""
        if isinstance(caption_obj, dict):
            caption_text = caption_obj.get("text", "")
        elif isinstance(caption_obj, str):
            caption_text = caption_obj

        hashtags = re.findall(r"#(\w+)", caption_text)

        view_count = item.get("view_count") or item.get("play_count") or 0
        like_count = (
            item.get("like_count")
            or item.get("likes", {}).get("count", 0)
            or 0
        )
        comment_count = (
            item.get("comment_count")
            or item.get("comments", {}).get("count", 0)
            or 0
        )

        clips = item.get("clips_metadata", {}) or {}
        orig_sound = clips.get("original_sound_info", {}) or {}
        audio_name = orig_sound.get("original_audio_title", "")
        audio_id = str(orig_sound.get("audio_asset_id", "") or orig_sound.get("xma_asset_id", "") or "")
        if not audio_name or not audio_id:
            music = item.get("music_metadata", {}) or {}
            music_asset = (music.get("music_info", {}) or {}).get("music_asset_info", {}) or {}
            audio_name = audio_name or music_asset.get("title", "")
            audio_id = audio_id or str(music_asset.get("audio_cluster_id", "") or "")

        taken_at = item.get("taken_at", 0)

        return {
            "reel_url": reel_url,
            "username": username,
            "full_name": full_name,
            "followers": followers,
            "profile_pic": profile_pic,
            "caption": caption_text,
            "hashtags": hashtags,
            "views": view_count,
            "likes": like_count,
            "comments": comment_count,
            "audio": audio_name,
            "audio_id": audio_id,
            "posted_at": taken_at,
        }
    except Exception:
        return None


async def scrape_reels_feed(username, password, max_reels=200, progress_callback=None,
                            scraped_by="", batch_id="", session_id=None):
    reels_collected = []
    seen_urls = set()

    async with async_playwright() as p:
        context, page = await _launch_context(p, username)

        captured_responses = []
        resp_counter = {"n": 0}

        async def handle_response(response):
            url = response.url
            if not any(kw in url for kw in ["/api/", "graphql", "clips", "reels", "feed"]):
                return
            ct = response.headers.get("content-type", "")
            if "json" not in ct:
                return
            try:
                body = await response.json()
            except Exception:
                return
            captured_responses.append(body)
            resp_counter["n"] += 1
            debug_dump_response(url, body, resp_counter["n"])

        page.on("response", handle_response)
        debug_log(f"=== New scrape run, max_reels={max_reels} ===")

        if progress_callback:
            progress_callback("Opening Instagram...")

        logged_in = await ensure_logged_in(page, context, username, progress_callback, session_id=session_id)
        if not logged_in:
            await context.close()
            return None, "Login was not completed in time. Please run again and log in."

        if progress_callback:
            progress_callback("Logged in. Navigating to Reels feed...")

        await safe_goto(page, f"{INSTAGRAM_URL}/reels/")

        # Wait for the reels player to mount — look for the first video element.
        try:
            await page.wait_for_selector("video", timeout=10000)
        except Exception:
            debug_log("Reels page: no <video> found within 10s — continuing anyway.")
        await page.wait_for_timeout(2000)

        # Flush any responses captured during login/home-page load so we only
        # collect reels from the dedicated /reels/ feed going forward.
        captured_responses.clear()
        seen_urls.clear()

        # Click the first reel to give the player keyboard focus.
        # Without this, ArrowDown scrolls the page body instead of advancing reels.
        try:
            first_video = page.locator("video").first
            await first_video.click(timeout=5000)
            debug_log("Clicked first reel — player has focus.")
        except Exception as e:
            debug_log(f"Could not click first reel ({e}); ArrowDown may scroll page body instead.")
        await page.wait_for_timeout(1500)

        # Fresh flush — capture the first reel that was already playing before we start scrolling.
        for resp_data in list(captured_responses):
            new_reels = await extract_reel_data_from_response(resp_data)
            for r in new_reels:
                if r["reel_url"] not in seen_urls:
                    seen_urls.add(r["reel_url"])
                    reels_collected.append(r)
        captured_responses.clear()

        if progress_callback:
            progress_callback(f"Reels feed open. Starting scroll ({max_reels} target)...")

        scroll_attempts = 0
        max_scrolls = 60

        while len(reels_collected) < max_reels and scroll_attempts < max_scrolls:
            await page.keyboard.press("ArrowDown")
            await page.wait_for_timeout(1800)

            for resp_data in captured_responses:
                new_reels = await extract_reel_data_from_response(resp_data)
                for r in new_reels:
                    if r["reel_url"] not in seen_urls:
                        seen_urls.add(r["reel_url"])
                        reels_collected.append(r)
            captured_responses.clear()

            if progress_callback:
                progress_callback(f"Collected {len(reels_collected)} reels so far...")

            scroll_attempts += 1

            if scroll_attempts % 5 == 0:
                await page.wait_for_timeout(1000)

        # Final flush
        for resp_data in captured_responses:
            new_reels = await extract_reel_data_from_response(resp_data)
            for r in new_reels:
                if r["reel_url"] not in seen_urls:
                    seen_urls.add(r["reel_url"])
                    reels_collected.append(r)

        # Enrich — fetch follower count, bio, email for each unique creator
        if reels_collected:
            if progress_callback:
                progress_callback(f"📊 Enriching {len(reels_collected)} creators (followers, email, bio)...")
            await enrich_creators(page, reels_collected, progress_callback)

        await context.close()

    saved_path = save_reels_csv(reels_collected, scraped_by=scraped_by,
                                batch_id=batch_id, scraped_from=username, push_db=False)
    debug_log(
        f"Run finished: {len(reels_collected)} reels parsed, "
        f"{resp_counter['n']} JSON responses dumped to debug/, "
        f"CSV: {saved_path}"
    )
    return reels_collected, None


_SC_ALPHABET = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_"


def shortcode_to_pk(shortcode):
    """Convert an Instagram shortcode (e.g. from /reel/DUnTahYAXWW/) to its numeric
    media id (pk), so we can hit /api/v1/media/<pk>/info/ directly."""
    if not shortcode:
        return None
    pk = 0
    for ch in shortcode:
        if ch not in _SC_ALPHABET:
            return None
        pk = pk * 64 + _SC_ALPHABET.index(ch)
    return pk if pk > 0 else None


async def fetch_profile_info(page, username):
    """Fetch a creator's profile via the web_profile_info API: follower count,
    category, bio, contact email, external link. Returns a dict or None.
    Retries once on 429 after a longer backoff."""
    for attempt in range(2):
        try:
            resp = await page.request.get(
                f"{INSTAGRAM_URL}/api/v1/users/web_profile_info/?username={username}",
                headers={"x-ig-app-id": "936619743392459",
                         "referer": f"{INSTAGRAM_URL}/{username}/"},
            )
            if resp.status == 429:
                if attempt == 0:
                    debug_log(f"profile {username} 429 — backing off 20s before retry")
                    await page.wait_for_timeout(20000)
                    continue
                debug_log(f"profile {username} 429 on retry — skipping")
                return None
            if not resp.ok:
                debug_log(f"profile {username} HTTP {resp.status}")
                return None
            u = ((await resp.json()).get("data") or {}).get("user") or {}
            followers = u.get("follower_count")
            if followers is None:
                followers = (u.get("edge_followed_by") or {}).get("count", 0)
            return {
                "followers": int(followers or 0),
                "category": u.get("category_name") or u.get("category") or "",
                "bio": (u.get("biography") or "").replace("\n", " ")[:300],
                "contact_email": (u.get("business_email") or u.get("public_email") or ""),
                "external_url": u.get("external_url") or "",
                "is_verified": bool(u.get("is_verified", False)),
            }
        except Exception as e:
            debug_log(f"profile {username} error: {e}")
            return None
    return None


async def enrich_creators(page, reels, progress_callback=None, known_profiles=None):
    """For each UNIQUE creator in `reels`, fetch profile info once and merge it into
    all of that creator's reels (followers, category, bio, email).
    If known_profiles dict is passed, creators already in DB are filled from cache
    — no API call needed, much faster."""
    known_profiles = known_profiles or {}
    by_user = {}
    for r in reels:
        u = r.get("username")
        if u:
            by_user.setdefault(u, []).append(r)

    cached, to_fetch = [], []
    for uname in by_user:
        if uname in known_profiles:
            cached.append(uname)
        else:
            to_fetch.append(uname)

    # Fill known creators instantly from cache
    for uname in cached:
        for r in by_user[uname]:
            r.update(known_profiles[uname])

    debug_log(f"Enrichment: {len(cached)} from cache, {len(to_fetch)} need API calls.")
    if progress_callback and cached:
        progress_callback(f"{len(cached)} creators loaded from DB cache, fetching {len(to_fetch)} new ones...")

    done = 0
    for uname in to_fetch:
        try:
            info = await fetch_profile_info(page, uname)
            if info:
                for r in by_user[uname]:
                    r.update(info)
            done += 1
            if progress_callback and done % 5 == 0:
                progress_callback(f"Enriching creators {done}/{len(to_fetch)} — followers + email...")
            # 1.5-2.5s random delay — keeps us well under Instagram's rate limit.
            await page.wait_for_timeout(1500 + random.randint(0, 1000))
            # Extra pause every 10 creators to stay under sustained rate limits.
            if done % 10 == 0:
                await page.wait_for_timeout(5000)
        except Exception as e:
            debug_log(f"Enrichment stopped early at {uname}: {e}")
            if progress_callback:
                progress_callback(f"Enrichment stopped at {done}/{len(to_fetch)} — saving what we have...")
            break  # Save partial results instead of crashing

    debug_log(f"Enrichment done: {len(cached)} cached + {len(to_fetch)} fetched.")
    return reels


async def fetch_creator(page, username):
    """One web_profile_info call → a full creator record: pk (for chaining),
    followers/category/email/bio + a representative recent reel (url, likes, caption)."""
    try:
        resp = await page.request.get(
            f"{INSTAGRAM_URL}/api/v1/users/web_profile_info/?username={username}",
            headers={"x-ig-app-id": "936619743392459",
                     "referer": f"{INSTAGRAM_URL}/{username}/"},
        )
        if not resp.ok:
            debug_log(f"creator {username} HTTP {resp.status}")
            return None
        u = ((await resp.json()).get("data") or {}).get("user") or {}
    except Exception as e:
        debug_log(f"creator {username} error: {e}")
        return None

    followers = u.get("follower_count")
    if followers is None:
        followers = (u.get("edge_followed_by") or {}).get("count", 0)

    # Pick a representative recent reel from the profile's timeline (prefer video).
    reel_url = f"{INSTAGRAM_URL}/{username}/"
    likes, caption = 0, ""
    edges = (u.get("edge_owner_to_timeline_media") or {}).get("edges", [])
    chosen = None
    for e in edges:
        node = (e or {}).get("node", {})
        if node.get("is_video") or node.get("product_type") == "clips":
            chosen = node
            break
    if chosen is None and edges:
        chosen = edges[0].get("node", {})
    if chosen:
        code = chosen.get("shortcode") or chosen.get("code")
        if code:
            reel_url = f"{INSTAGRAM_URL}/reel/{code}/"
        likes = ((chosen.get("edge_media_preview_like") or {}).get("count")
                 or (chosen.get("edge_liked_by") or {}).get("count")
                 or chosen.get("like_count") or 0)
        cap_edges = (chosen.get("edge_media_to_caption") or {}).get("edges", [])
        if cap_edges:
            caption = (cap_edges[0].get("node", {}) or {}).get("text", "") or ""

    return {
        "pk": u.get("id") or u.get("pk"),
        "reel_url": reel_url,
        "username": username,
        "full_name": u.get("full_name", "") or "",
        "followers": int(followers or 0),
        "category": u.get("category_name") or u.get("category") or "",
        "bio": (u.get("biography") or "").replace("\n", " ")[:300],
        "contact_email": (u.get("business_email") or u.get("public_email") or ""),
        "external_url": u.get("external_url") or "",
        "is_verified": bool(u.get("is_verified", False)),
        "likes": int(likes or 0),
        "comments": 0,
        "views": 0,
        "caption": caption,
        "hashtags": re.findall(r"#(\w+)", caption),
        "audio": "",
    }


def _extract_usernames_from_data(data, result, seen, depth=0):
    """Recursively pull username strings from any API response shape."""
    if depth > 10 or len(result) > 200:
        return
    if isinstance(data, dict):
        uname = data.get("username", "")
        if uname and isinstance(uname, str) and uname not in seen:
            seen.add(uname)
            result.append(uname)
        for v in data.values():
            _extract_usernames_from_data(v, result, seen, depth + 1)
    elif isinstance(data, list):
        for item in data:
            _extract_usernames_from_data(item, result, seen, depth + 1)


async def fetch_similar_from_profile(page, username, pk):
    """
    3-method approach to get similar/suggested accounts for a seed creator.

    Method 1 — Chaining API (with profile page as referer, not "/" like before)
    Method 2 — Visit profile page, capture all JSON API responses fired on load
    Method 3 — DOM: extract profile links from the page (Similar accounts section)

    Returns list of usernames (may be empty if all methods fail).
    """
    similar = []
    seen = {username.lower()}

    # ── Method 1: Chaining API with profile-page referer ─────────────────────
    # Previous attempt used referer="/" which may have been why it returned 400.
    if pk:
        try:
            resp = await page.request.get(
                f"{INSTAGRAM_URL}/api/v1/discover/chaining/?target_id={pk}",
                headers={"x-ig-app-id": "936619743392459",
                         "referer": f"{INSTAGRAM_URL}/{username}/"},
            )
            debug_log(f"chaining @{username} pk={pk} → HTTP {resp.status}")
            if resp.ok:
                body = await resp.json()
                debug_dump_response(f"chaining_{username}", body, 0)
                for u in (body.get("users") or []):
                    uname = u.get("username", "")
                    if uname and uname.lower() not in seen:
                        seen.add(uname.lower())
                        similar.append(uname)
                debug_log(f"Chaining: {len(similar)} similar accounts for @{username}")
        except Exception as e:
            debug_log(f"Chaining error @{username}: {e}")

    if similar:
        return similar

    # ── Method 2: Visit profile page, capture API responses ──────────────────
    captured_users = []
    cap_seen = set(seen)

    async def _on_response(response):
        if "json" not in response.headers.get("content-type", ""):
            return
        try:
            data = await response.json()
            _extract_usernames_from_data(data, captured_users, cap_seen)
        except Exception:
            pass

    page.on("response", _on_response)
    await safe_goto(page, f"{INSTAGRAM_URL}/{username}/")
    await page.wait_for_timeout(4000)
    # Scroll to trigger "Similar accounts" section loading
    for _ in range(3):
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight / 3)")
        await page.wait_for_timeout(1000)
    page.remove_listener("response", _on_response)

    # Filter out the seed themselves and obviously non-creator system accounts
    SYSTEM_ACCOUNTS = {"instagram", "meta", "creators", "shop", "reels", "explore"}
    for uname in captured_users:
        if uname.lower() not in seen and uname.lower() not in SYSTEM_ACCOUNTS:
            seen.add(uname.lower())
            similar.append(uname)

    if similar:
        debug_log(f"Profile page API capture: {len(similar)} usernames for @{username}")
        return similar

    # ── Method 3: DOM — extract all /username/ links from the page ───────────
    # Page is already on the profile from Method 2 — no extra navigation needed.
    SKIP_PATHS = {
        "explore", "reels", "stories", "direct", "accounts", "p", "reel",
        "tv", "about", "help", "press", "api", "privacy", "legal",
        "developers", "blog", "jobs", "directory",
        # Instagram system accounts
        "instagram", "meta", "creators", "shop", "create",
        username.lower(), "",
    }
    try:
        hrefs = await page.eval_on_selector_all(
            'a[href^="/"]',
            "els => els.map(e => e.getAttribute('href'))"
        )
        for href in (hrefs or []):
            m = re.match(r'^/([A-Za-z0-9._]{2,30})/?$', href or "")
            if m:
                uname = m.group(1)
                if uname.lower() not in seen and uname.lower() not in SKIP_PATHS:
                    seen.add(uname.lower())
                    similar.append(uname)
    except Exception as e:
        debug_log(f"DOM extraction error @{username}: {e}")

    debug_log(f"DOM method: {len(similar)} profile links for @{username}")
    return similar


async def fetch_seed_signals(page, username):
    """Analyse a seed creator's recent reels to extract discovery signals:
      - pk: numeric user ID
      - related: edge_related_profiles (usually empty now)
      - hashtags: top hashtags from their captions (may be generic/empty)
      - audio_ids: unique audio IDs from their reels (for audio-page search)
      - audio_names: display names matching audio_ids (for UI)
    Returns a dict with all of the above."""
    from collections import Counter
    out = {"related": [], "hashtags": [], "pk": None, "audio_ids": [], "audio_names": []}

    # Step 1: web_profile_info → pk (+ related profiles if present).
    try:
        resp = await page.request.get(
            f"{INSTAGRAM_URL}/api/v1/users/web_profile_info/?username={username}",
            headers={"x-ig-app-id": "936619743392459",
                     "referer": f"{INSTAGRAM_URL}/{username}/"},
        )
        if resp.ok:
            body = await resp.json()
            try:
                debug_dump_response(f"web_profile_info_{username}", body, 0)
            except Exception:
                pass
            u = (body.get("data") or {}).get("user") or {}
            out["pk"] = u.get("id") or u.get("fbid")
            for e in (u.get("edge_related_profiles") or {}).get("edges", []):
                n = ((e or {}).get("node") or {}).get("username")
                if n:
                    out["related"].append(n)
        else:
            debug_log(f"seed web_profile_info {username} HTTP {resp.status}")
    except Exception as e:
        debug_log(f"seed web_profile_info {username} error: {e}")

    # Step 2: the seed's recent posts (real captions) → hashtags.
    if out["pk"]:
        try:
            r2 = await page.request.get(
                f"{INSTAGRAM_URL}/api/v1/feed/user/{out['pk']}/?count=12",
                headers={"x-ig-app-id": "936619743392459",
                         "referer": f"{INSTAGRAM_URL}/{username}/"},
            )
            if r2.ok:
                data2 = await r2.json()
                try:
                    debug_dump_response(f"feed_user_{username}", data2, 0)
                except Exception:
                    pass
                cnt = Counter()
                seen_audio = set()
                for item in data2.get("items", []) or []:
                    # Hashtags from caption
                    cap = item.get("caption") or {}
                    txt = cap.get("text", "") if isinstance(cap, dict) else ""
                    for h in re.findall(r"#(\w+)", txt):
                        cnt[h.lower()] += 1

                    # Audio IDs — try clips_metadata first, then music_metadata
                    clips = item.get("clips_metadata", {}) or {}
                    orig = clips.get("original_sound_info", {}) or {}
                    a_id = str(orig.get("audio_asset_id", "") or orig.get("xma_asset_id", "") or "")
                    a_name = orig.get("original_audio_title", "")
                    if not a_id:
                        music = item.get("music_metadata", {}) or {}
                        masset = (music.get("music_info", {}) or {}).get("music_asset_info", {}) or {}
                        a_id = str(masset.get("audio_cluster_id", "") or "")
                        a_name = masset.get("title", "") or a_name
                    if a_id and a_id not in seen_audio:
                        seen_audio.add(a_id)
                        out["audio_ids"].append(a_id)
                        out["audio_names"].append(a_name or f"Audio {a_id[:8]}")

                out["hashtags"] = [h for h, _ in cnt.most_common(10)]
            else:
                debug_log(f"seed feed/user {username} HTTP {r2.status}")
        except Exception as e:
            debug_log(f"seed feed/user {username} error: {e}")

    debug_log(f"seed @{username}: pk={out['pk']}, {len(out['related'])} related, "
              f"hashtags={out['hashtags'][:8]}")
    return out


def _detect_styles_from_hashtags(hashtags):
    """
    Given a list of hashtag strings from a seed's reels, detect which creator
    styles they match using the CREATOR_STYLE_TAGS detect lists.
    Returns a list of matched style names (e.g. ["😂 Stand-up / Comedy", "🎤 Direct to Camera"]).
    """
    from filters import CREATOR_STYLE_TAGS
    tag_set = {h.lower().replace("#", "").replace("_", "") for h in hashtags}
    detected = []
    for style, data in CREATOR_STYLE_TAGS.items():
        detect_tags = (
            {t.replace("_", "") for t in data.get("detect", [])}
            if isinstance(data, dict)
            else {t.replace("_", "") for t in data}
        )
        if tag_set & detect_tags:
            detected.append(style)
    return detected


async def scrape_seed_creators(username, seeds, max_creators=100, progress_callback=None,
                               scraped_by="", batch_id="", enrich=True, known_profiles=None,
                               mode="hashtags", session_id=None, depth=1):
    """
    Seed expansion via content analysis — mirrors what a human researcher would do:

    PHASE 1 — Analyze the seed
      • Fetch seed's last 12 reels via API (fast, no browser scroll needed)
      • Extract all hashtags from their captions
      • Detect their creator style(s) using CREATOR_STYLE_TAGS
      • Build search tag list = seed's own hashtags + style-specific search hashtags

    PHASE 2 — Search for similar creators
      • Run hashtag search on the combined tag list (same proven engine as Hashtag Search tab)
      • Each hashtag page on Explore shows creators posting in that niche

    PHASE 3 — Enrich
      • Fetch follower counts, bio, contact email for each discovered creator

    Returns (reels, signals_dict) where signals_dict contains detected styles + search tags
    so the app can display what was found.
    """
    from collections import Counter
    from filters import get_style_search_hashtags

    seed_names = [s.strip().lstrip("@") for s in seeds if s and s.strip()]
    if not seed_names:
        return None, "No seed creators provided."

    async with async_playwright() as p:
        context, page = await _launch_context(p, username)
        debug_log(f"=== Seed expansion: {seed_names}, max={max_creators} ===")

        logged_in = await ensure_logged_in(page, context, username, progress_callback, session_id=session_id)
        if not logged_in:
            await context.close()
            return None, "Login was not completed in time. Please run again and log in."

        # ── PHASE 1: Analyze each seed's content ──────────────────────────────
        tag_counter = Counter()
        all_detected_styles = []
        all_audio_ids, all_audio_names = [], []
        seen_audio = set()

        for seed in seed_names:
            if progress_callback:
                progress_callback(f"🔍 Analyzing @{seed}'s recent reels...")

            sig = await fetch_seed_signals(page, seed)
            seed_hashtags = sig.get("hashtags", [])

            for t in seed_hashtags:
                tag_counter[t] += 1

            styles = _detect_styles_from_hashtags(seed_hashtags)
            for s in styles:
                if s not in all_detected_styles:
                    all_detected_styles.append(s)

            # Collect unique audio IDs across all seeds
            for a_id, a_name in zip(sig.get("audio_ids", []), sig.get("audio_names", [])):
                if a_id and a_id not in seen_audio:
                    seen_audio.add(a_id)
                    all_audio_ids.append(a_id)
                    all_audio_names.append(a_name)

            debug_log(f"@{seed}: hashtags={seed_hashtags[:5]}, "
                      f"styles={styles}, audio_ids={sig.get('audio_ids', [])[:3]}")
            if progress_callback:
                style_str = ", ".join(styles) if styles else "style not detected"
                audio_count = len(sig.get("audio_ids", []))
                progress_callback(
                    f"✅ @{seed} → {style_str} | "
                    f"{len(seed_hashtags)} hashtags | {audio_count} audio tracks found"
                )
            await page.wait_for_timeout(300)

        # ── Build hashtag search list ──────────────────────────────────────────
        GENERIC_TAGS = {
            "reels", "reelsindia", "reelsinstagram", "reelsviral", "reelitfeelit",
            "instagram", "instagramreels", "instareels", "viral", "trending",
            "fyp", "foryou", "foryoupage", "explorepage", "explore",
            "instagood", "love", "like", "follow", "share", "video",
            "india", "indianreels", "desi", "bharat",
        }
        seed_top_tags = [
            t for t, _ in tag_counter.most_common(20)
            if t.lower() not in GENERIC_TAGS and len(t) > 3
        ][:6]
        style_search_tags = get_style_search_hashtags(all_detected_styles)
        combined_tags = list(dict.fromkeys(seed_top_tags + style_search_tags))[:12]

        debug_log(f"Mode={mode} | tags={combined_tags} | audio_ids={all_audio_ids[:4]}")

        # ── Validate we have something to search with ──────────────────────────
        has_tags = bool(combined_tags)
        has_audio = bool(all_audio_ids)
        if mode in ("hashtags", "both") and not has_tags and not has_audio:
            await context.close()
            return None, (
                "Couldn't find hashtags or audio from those seeds — "
                "they may be private or have no recent reels. Try other seeds."
            )
        if mode == "hashtags" and not has_tags:
            await context.close()
            return None, (
                "No usable hashtags found from those seeds. "
                "Try switching to Audio mode — the seed may not use hashtags."
            )
        if mode == "audio" and not has_audio:
            await context.close()
            return None, (
                "No audio tracks found from those seeds. "
                "Try switching to Hashtags mode."
            )

        seed_label = ", ".join(f"@{s}" for s in seed_names)
        raw_reels = []

        # ── PHASE 2: Search ────────────────────────────────────────────────────
        if mode in ("hashtags", "both") and combined_tags:
            if progress_callback:
                progress_callback(
                    f"🏷️ Searching hashtags: {', '.join('#'+t for t in combined_tags)}"
                )
            tag_reels = await _harvest_tag_reels(
                page, combined_tags, max_creators * 2, progress_callback
            )
            raw_reels.extend(tag_reels)

        if mode in ("audio", "both") and all_audio_ids:
            if progress_callback:
                progress_callback(
                    f"🎵 Searching {len(all_audio_ids)} audio tracks from seed's reels..."
                )
            audio_reels = await _harvest_audio_reels(
                page, all_audio_ids, all_audio_names,
                max_creators * 2, progress_callback
            )
            # Dedup by reel_url across both sources
            existing_urls = {r.get("reel_url") for r in raw_reels}
            raw_reels.extend(r for r in audio_reels if r.get("reel_url") not in existing_urls)

        # ── Style-match filter (hashtag mode only — audio results are already on-niche) ──
        if all_detected_styles and mode == "hashtags":
            from filters import CREATOR_STYLE_TAGS
            style_detect_tags = set()
            for style in all_detected_styles:
                data = CREATOR_STYLE_TAGS.get(style, {})
                tags = data.get("detect", data) if isinstance(data, dict) else data
                style_detect_tags.update(t.replace("_", "").lower() for t in tags)
            style_detect_tags.update(t.lower() for t in seed_top_tags)

            reels_collected = []
            for r in raw_reels:
                reel_tags = {
                    h.lower().replace("#", "").replace("_", "")
                    for h in r.get("hashtags", [])
                }
                if reel_tags & style_detect_tags:
                    reels_collected.append(r)
            debug_log(f"Style filter: {len(raw_reels)} → {len(reels_collected)} kept")
        else:
            reels_collected = raw_reels

        # ── PHASE 2b: Deep expand via Instagram's "similar accounts" (chaining) ──
        # ADDITIVE + GUARDED: if the chaining endpoint is restricted, this adds
        # nothing and the hashtag results above are completely unaffected.
        if depth and int(depth) >= 1:
            try:
                chain_seen = set(seed_names) | {r.get("username") for r in reels_collected}
                frontier = list(seed_names)
                chain_added = 0
                for _level in range(int(depth)):
                    if len(reels_collected) >= max_creators:
                        break
                    next_frontier = []
                    for su in frontier:
                        if len(reels_collected) >= max_creators:
                            break
                        base = await fetch_creator(page, su)
                        pk = base.get("pk") if base else None
                        if not pk:
                            continue
                        similar = await fetch_similar_from_profile(page, su, pk)
                        if progress_callback and similar:
                            progress_callback(f"🔗 @{su} → {len(similar)} similar creators")
                        for su2 in similar[:15]:
                            if su2 in chain_seen:
                                continue
                            chain_seen.add(su2)
                            next_frontier.append(su2)
                            crec = await fetch_creator(page, su2)
                            if crec and crec.get("username"):
                                reels_collected.append(crec)
                                chain_added += 1
                            if len(reels_collected) >= max_creators:
                                break
                        await page.wait_for_timeout(400)
                    frontier = next_frontier[:5]   # only chain from top 5 next level
                    if not frontier:
                        break
                debug_log(f"Chaining deep-expand: +{chain_added} creators (depth={depth})")
                if progress_callback:
                    progress_callback(f"🔗 Deep expand: +{chain_added} similar creators")
            except Exception as e:
                debug_log(f"Chaining deep-expand failed (ignored): {e}")

        for r in reels_collected:
            r.setdefault("source_seed", seed_label)
            r["detected_styles"] = all_detected_styles
            r["search_mode"] = mode

        # ── PHASE 3: Enrich ───────────────────────────────────────────────────
        if enrich and reels_collected:
            if progress_callback:
                progress_callback("📊 Fetching follower counts + contact emails...")
            await enrich_creators(page, reels_collected, progress_callback,
                                  known_profiles=known_profiles or {})

        await context.close()

    signals = {
        "detected_styles": all_detected_styles,
        "seed_hashtags": seed_top_tags,
        "style_hashtags": style_search_tags[:8],
        "search_tags": combined_tags,
        "audio_ids": all_audio_ids,
        "audio_names": all_audio_names,
        "mode": mode,
    }

    label = "like " + ", ".join(f"@{s}" for s in seed_names)
    save_reels_csv(reels_collected, scraped_by=scraped_by, batch_id=batch_id,
                   scraped_from=label, push_db=False)
    debug_log(f"Seed expansion finished: {len(reels_collected)} creators.")
    return reels_collected, signals


def _href_code(href):
    m = re.search(r"/(p|reel|tv)/([^/?#]+)", href or "")
    return m.group(2) if m else None


async def _harvest_tag_reels(page, tags, max_reels, progress_callback=None, max_scrolls=40):
    """Simple original: PHASE 1 harvest reel/post links from each tag's grid, then
    PHASE 2 fetch each via the media-info API. Works off the browser's logged-in
    session. Used by hashtag search + seed expansion."""
    reels, seen_urls, seen_codes, post_links = [], set(), set(), []
    # Sanitize tags: a hashtag is ONE token — split any that contain spaces/commas,
    # strip '#'. Prevents the "two tags became one keyword search" bug.
    clean_tags = []
    for t in tags:
        for part in str(t).replace(",", " ").split():
            part = part.strip().lstrip("#")
            if part and part not in clean_tags:
                clean_tags.append(part)
    tags = clean_tags
    per_tag = max(15, max_reels // max(1, len(tags) or 1))

    # PHASE 1 — harvest links from each tag's grid.
    for tag in tags:
        if progress_callback:
            progress_callback(f"Opening #{tag}...")
        await safe_goto(page, f"{INSTAGRAM_URL}/explore/tags/{tag}/")
        await page.wait_for_timeout(5000)
        tag_links, scrolls = 0, 0
        while tag_links < per_tag and scrolls < max_scrolls and len(post_links) < max_reels:
            try:
                hrefs = await page.eval_on_selector_all(
                    'a[href*="/reel/"], a[href*="/p/"]',
                    "els => els.map(e => e.getAttribute('href'))")
            except Exception:
                hrefs = []
            for href in hrefs:
                c = _href_code(href)
                if c and c not in seen_codes:
                    seen_codes.add(c)
                    post_links.append((href, tag))
                    tag_links += 1
            if progress_callback:
                progress_callback(f"#{tag}: found {tag_links} links · total {len(post_links)}")
            try:
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            except Exception:
                pass
            await page.keyboard.press("End")
            await page.wait_for_timeout(2000)
            scrolls += 1
        debug_log(f"#{tag}: harvested {tag_links} links")

    # PHASE 2 — fetch each post via the media-info API.
    targets = post_links[:max_reels]
    fails = 0
    for idx, (href, tag) in enumerate(targets, 1):
        pk = shortcode_to_pk(_href_code(href))
        if not pk:
            continue
        try:
            resp = await page.request.get(
                f"{INSTAGRAM_URL}/api/v1/media/{pk}/info/",
                headers={"x-ig-app-id": "936619743392459",
                         "referer": INSTAGRAM_URL + (href if href.startswith("/") else "/")},
            )
            if resp.ok:
                for r in await extract_reel_data_from_response(await resp.json()):
                    if r["reel_url"] not in seen_urls:
                        seen_urls.add(r["reel_url"])
                        r["source_hashtag"] = tag
                        reels.append(r)
            else:
                fails += 1
                debug_log(f"media info pk={pk} HTTP {resp.status}")
        except Exception as e:
            fails += 1
            debug_log(f"media info error {_href_code(href)}: {e}")
        if progress_callback and idx % 5 == 0:
            progress_callback(f"Fetched {idx}/{len(targets)} · {len(reels)} creators")
        await page.wait_for_timeout(450)
    debug_log(f"Phase 2 done: {len(reels)} creators, {fails} failures.")
    return reels



async def get_creator_hashtags(page, username, want=8):
    """Read a seed creator's recent reels and return their most-used hashtags."""
    from collections import Counter
    try:
        resp = await page.request.get(
            f"{INSTAGRAM_URL}/api/v1/users/web_profile_info/?username={username}",
            headers={"x-ig-app-id": "936619743392459",
                     "referer": f"{INSTAGRAM_URL}/{username}/"},
        )
        if not resp.ok:
            debug_log(f"seed hashtags {username} HTTP {resp.status}")
            return []
        u = ((await resp.json()).get("data") or {}).get("user") or {}
    except Exception as e:
        debug_log(f"seed hashtags {username} error: {e}")
        return []
    cnt = Counter()
    for e in (u.get("edge_owner_to_timeline_media") or {}).get("edges", [])[:12]:
        node = (e or {}).get("node", {})
        cap_edges = (node.get("edge_media_to_caption") or {}).get("edges", [])
        text = cap_edges[0].get("node", {}).get("text", "") if cap_edges else ""
        for h in re.findall(r"#(\w+)", text):
            cnt[h.lower()] += 1
    tags = [h for h, _ in cnt.most_common(want)]
    debug_log(f"seed @{username} top hashtags: {tags}")
    return tags


async def _harvest_audio_reels(page, audio_ids, audio_names, max_reels,
                               progress_callback=None, max_scrolls=40):
    """
    Same harvest logic as _harvest_tag_reels but visits /reels/audio/{id}/ pages.
    audio_ids: list of Instagram audio cluster IDs
    audio_names: parallel list of display names (for progress messages)
    Returns a list of reel records (same format as _harvest_tag_reels).
    """
    reels, seen_urls, seen_codes, post_links = [], set(), set(), []
    per_audio = max(15, max_reels // max(1, len(audio_ids)))

    for a_id, a_name in zip(audio_ids, audio_names):
        if progress_callback:
            progress_callback(f"🎵 Searching audio: {a_name or a_id[:12]}...")
        await safe_goto(page, f"{INSTAGRAM_URL}/reels/audio/{a_id}/")
        await page.wait_for_timeout(5000)
        audio_links, scrolls = 0, 0
        while audio_links < per_audio and scrolls < max_scrolls and len(post_links) < max_reels:
            try:
                hrefs = await page.eval_on_selector_all(
                    'a[href*="/reel/"], a[href*="/p/"]',
                    "els => els.map(e => e.getAttribute('href'))")
            except Exception:
                hrefs = []
            for href in hrefs:
                c = _href_code(href)
                if c and c not in seen_codes:
                    seen_codes.add(c)
                    post_links.append((href, f"audio:{a_id}"))
                    audio_links += 1
            await page.evaluate("window.scrollBy(0, 1200)")
            await page.wait_for_timeout(2000)
            scrolls += 1
        debug_log(f"audio {a_id}: {audio_links} links after {scrolls} scrolls")

    if not post_links:
        return []

    if progress_callback:
        progress_callback(f"Found {len(post_links)} reels from audio pages — fetching details...")

    # Phase 2: same as _harvest_tag_reels — fetch media info for each link
    for href, source in post_links:
        code = _href_code(href)
        if not code:
            continue
        try:
            pk = shortcode_to_pk(code)
            resp = await page.request.get(
                f"{INSTAGRAM_URL}/api/v1/media/{pk}/info/",
                headers={"x-ig-app-id": "936619743392459",
                         "referer": INSTAGRAM_URL + "/"},
            )
            if not resp.ok:
                continue
            body = await resp.json()
            item = _deep_find_media(body)
            if not item:
                continue
            rec = parse_reel_item(item)
            if rec and rec.get("reel_url") and rec["reel_url"] not in seen_urls:
                seen_urls.add(rec["reel_url"])
                reels.append(rec)
        except Exception as e:
            debug_log(f"audio reel fetch error {href}: {e}")
        await page.wait_for_timeout(300)

    debug_log(f"Audio harvest: {len(reels)} reels from {len(audio_ids)} audio pages")
    return reels


# Map a hashtag to a language by the language word inside it — reliable because
# the user searches language-specific tags (e.g. #tamilskit → Tamil).
_LANG_BY_TAG = [
    ("tamil", "Tamil"),
    ("bangla", "Bengali"), ("bengali", "Bengali"),
    ("telugu", "Telugu"),
    ("kannada", "Kannada"),
    ("malayalam", "Malayalam"),
    ("marathi", "Marathi"),
    ("punjabi", "Punjabi"), ("punjab", "Punjabi"),
    ("gujarati", "Gujarati"),
    ("bhojpuri", "Bhojpuri"),
    ("assamese", "Assamese"), ("assam", "Assamese"),
    ("odia", "Odia"), ("oriya", "Odia"),
    ("urdu", "Urdu"),
    ("hindi", "Hindi"),
]


def _language_from_tag(tag):
    t = (tag or "").lower()
    for key, lang in _LANG_BY_TAG:
        if key in t:
            return lang
    return ""


async def scrape_hashtags(username, hashtags, max_reels=200, progress_callback=None,
                          scraped_by="", batch_id="", enrich=True, known_profiles=None,
                          max_scrolls=40, session_id=None):
    """Scrape reels from one or more hashtag pages (explore/tags/<tag>/).
    Far more targeted than the home feed: pulls from a topic-defined pool.
    If enrich=True, also fetches each creator's followers/category/email."""
    reels_collected = []
    seen_urls = set()
    tags = [t.strip().lstrip("#") for t in hashtags if t and t.strip()]
    if not tags:
        return None, "No hashtags provided."

    async with async_playwright() as p:
        context, page = await _launch_context(p, username)
        debug_log(f"=== Hashtag scrape: {tags}, max_reels={max_reels} ===")

        logged_in = await ensure_logged_in(page, context, username, progress_callback, session_id=session_id)
        if not logged_in:
            await context.close()
            return None, "Login was not completed in time. Please run again and log in."

        reels_collected = await _harvest_tag_reels(page, tags, max_reels, progress_callback,
                                                   max_scrolls=max_scrolls)

        # Set language reliably from the hashtag that found each creator
        # (e.g. found via #tamilskit → Tamil). Overrides weak caption detection.
        for r in reels_collected:
            lang = _language_from_tag(r.get("source_hashtag", ""))
            if lang:
                r["detected_language"] = lang

        if enrich and reels_collected:
            if progress_callback:
                progress_callback("Fetching follower counts + contact emails...")
            await enrich_creators(page, reels_collected, progress_callback, known_profiles)

        await context.close()

    label = ", ".join(f"#{t}" for t in tags)
    saved_path = save_reels_csv(reels_collected, scraped_by=scraped_by,
                                batch_id=batch_id, scraped_from=label, push_db=False)
    debug_log(f"Hashtag run finished: {len(reels_collected)} reels from {label}, CSV: {saved_path}")
    return reels_collected, None


async def extract_reference_reel(username, password, reel_url, progress_callback=None,
                                 session_id=None):
    async with async_playwright() as p:
        context, page = await _launch_context(p, username)

        captured = []

        async def handle_response(response):
            url = response.url
            if not any(kw in url for kw in ["/api/", "graphql", "media", "clips", "reel"]):
                return
            if "json" not in response.headers.get("content-type", ""):
                return
            try:
                captured.append(await response.json())
            except Exception:
                pass

        page.on("response", handle_response)

        logged_in = await ensure_logged_in(page, context, username, progress_callback, session_id=session_id)
        if not logged_in:
            await context.close()
            return None, "Login was not completed in time. Please run again and log in."

        await safe_goto(page, reel_url)
        await page.wait_for_timeout(4000)

        reel_data = None
        for resp_data in captured:
            reels = await extract_reel_data_from_response(resp_data)
            if reels:
                reel_data = reels[0]
                break

        if not reel_data:
            try:
                caption_el = page.locator('div[data-testid="post-comment-root"] span, h1, div._a9zs span')
                caption_text = await caption_el.first.inner_text() if await caption_el.count() > 0 else ""
                hashtags = re.findall(r"#(\w+)", caption_text)
                reel_data = {
                    "reel_url": reel_url, "username": "",
                    "caption": caption_text, "hashtags": hashtags,
                    "views": 0, "likes": 0,
                }
            except Exception:
                pass

        await context.close()
        return reel_data, None


def run_scrape(username, password, max_reels, progress_callback=None, scraped_by="", session_id=None):
    # Each run gets a unique batch id so a teammate can isolate one scrape session.
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    return asyncio.run(scrape_reels_feed(
        username, password, max_reels, progress_callback,
        scraped_by=scraped_by, batch_id=batch_id, session_id=session_id,
    ))


def run_scrape_hashtags(username, hashtags, max_reels, progress_callback=None,
                        scraped_by="", enrich=True, known_profiles=None, max_scrolls=40,
                        session_id=None):
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    return asyncio.run(scrape_hashtags(
        username, hashtags, max_reels, progress_callback,
        scraped_by=scraped_by, batch_id=batch_id, enrich=enrich,
        known_profiles=known_profiles, max_scrolls=max_scrolls, session_id=session_id,
    ))


def run_scrape_seed(username, seeds, max_creators, progress_callback=None,
                    scraped_by="", known_profiles=None, mode="hashtags", session_id=None,
                    depth=1):
    """
    Returns (reels, signals, error_str).
    mode: "hashtags" | "audio" | "both"
    depth: how many "similar accounts" chaining levels to expand (0 = off, 1-2 typical)
    """
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    reels, result = asyncio.run(scrape_seed_creators(
        username, seeds, max_creators, progress_callback,
        scraped_by=scraped_by, batch_id=batch_id,
        known_profiles=known_profiles or {},
        mode=mode, session_id=session_id, depth=depth,
    ))
    if isinstance(result, str):
        return None, {}, result
    return reels, (result or {}), None


def run_extract_reference(username, password, reel_url):
    return asyncio.run(extract_reference_reel(username, password, reel_url))


async def scrape_from_reference_reel(username, reel_url, max_creators=100,
                                      progress_callback=None, scraped_by="",
                                      batch_id="", known_profiles=None, session_id=None):
    """
    Reference Reel flow:
    1. Open the reel URL, extract its hashtags + detected style
    2. Build search tag list (same logic as scrape_seed_creators)
    3. Run hashtag search on Explore with those tags
    Returns (reels, signals) — same shape as scrape_seed_creators.
    """
    from filters import get_style_search_hashtags

    GENERIC_TAGS = {
        "reels", "reelsindia", "reelsinstagram", "reelsviral", "reelitfeelit",
        "instagram", "instagramreels", "instareels", "viral", "trending",
        "fyp", "foryou", "foryoupage", "explorepage", "explore",
        "instagood", "love", "like", "follow", "share", "video",
        "india", "indianreels", "desi", "bharat",
    }

    async with async_playwright() as p:
        context, page = await _launch_context(p, username)

        logged_in = await ensure_logged_in(page, context, username, progress_callback, session_id=session_id)
        if not logged_in:
            await context.close()
            return None, "Login was not completed in time."

        # Step 1 — Extract reel data from the URL
        if progress_callback:
            progress_callback("🎬 Opening reference reel...")

        captured = []

        async def handle_response(response):
            if not any(kw in response.url for kw in ["/api/", "graphql", "media", "clips", "reel"]):
                return
            if "json" not in response.headers.get("content-type", ""):
                return
            try:
                captured.append(await response.json())
            except Exception:
                pass

        page.on("response", handle_response)
        await safe_goto(page, reel_url)
        await page.wait_for_timeout(4000)
        page.remove_listener("response", handle_response)

        reel_data = None
        for resp_data in captured:
            reels_found = await extract_reel_data_from_response(resp_data)
            if reels_found:
                reel_data = reels_found[0]
                break

        if not reel_data:
            # DOM fallback
            try:
                caption_el = page.locator('div[data-testid="post-comment-root"] span, h1, div._a9zs span')
                caption_text = await caption_el.first.inner_text() if await caption_el.count() > 0 else ""
                reel_data = {"hashtags": re.findall(r"#(\w+)", caption_text), "caption": caption_text}
            except Exception:
                reel_data = {"hashtags": [], "caption": ""}

        reel_hashtags = [h.lower() for h in (reel_data.get("hashtags") or [])]
        reel_username = reel_data.get("username", "")

        if progress_callback:
            progress_callback(f"✅ Reel extracted — {len(reel_hashtags)} hashtags found")

        # Step 2 — Build search tag list
        detected_styles = _detect_styles_from_hashtags(reel_hashtags)
        niche_tags = [t for t in reel_hashtags if t not in GENERIC_TAGS and len(t) > 3][:6]
        style_search_tags = get_style_search_hashtags(detected_styles)
        combined_tags = list(dict.fromkeys(niche_tags + style_search_tags))[:12]

        debug_log(f"Ref reel: styles={detected_styles}, tags={combined_tags}")

        if not combined_tags:
            await context.close()
            return None, (
                "No usable hashtags found in this reel. "
                "The reel may have no caption or only generic hashtags."
            )

        if progress_callback:
            progress_callback(f"🏷️ Searching: {', '.join('#'+t for t in combined_tags)}")

        # Step 3 — Hashtag search on Explore
        reels_collected = await _harvest_tag_reels(page, combined_tags, max_creators * 2, progress_callback)

        # Style-match filter
        if detected_styles:
            from filters import CREATOR_STYLE_TAGS
            style_detect_tags = set()
            for style in detected_styles:
                data = CREATOR_STYLE_TAGS.get(style, {})
                tags = data.get("detect", data) if isinstance(data, dict) else data
                style_detect_tags.update(t.replace("_", "").lower() for t in tags)
            style_detect_tags.update(t.lower() for t in niche_tags)
            reels_collected = [
                r for r in reels_collected
                if {h.lower().replace("#", "").replace("_", "") for h in r.get("hashtags", [])} & style_detect_tags
            ] or reels_collected  # keep all if filter removes everything

        if enrich := bool(reels_collected):
            if progress_callback:
                progress_callback("📊 Fetching follower counts + contact emails...")
            await enrich_creators(page, reels_collected, progress_callback, known_profiles or {})

        await context.close()

    signals = {
        "reel_username": reel_username,
        "reel_hashtags": reel_hashtags[:8],
        "detected_styles": detected_styles,
        "search_tags": combined_tags,
        "style_hashtags": style_search_tags[:6],
    }
    save_reels_csv(reels_collected, scraped_by=scraped_by, batch_id=batch_id,
                   scraped_from=f"ref_reel:{reel_url}", push_db=False)
    return reels_collected, signals


def run_scrape_reference_reel(username, reel_url, max_creators, progress_callback=None,
                               scraped_by="", known_profiles=None, session_id=None):
    """Returns (reels, signals, error_str)."""
    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    result = asyncio.run(scrape_from_reference_reel(
        username, reel_url, max_creators, progress_callback,
        scraped_by=scraped_by, batch_id=batch_id, known_profiles=known_profiles or {},
        session_id=session_id,
    ))
    if isinstance(result[1], str):   # error string
        return None, {}, result[1]
    return result[0], result[1], None


async def _open_training_browser_async(username, session_id=None):
    """Open the persistent browser for `username` so the user can train their feed.
    Blocks until the browser window is closed."""
    async with async_playwright() as p:
        context, page = await _launch_context(p, username)
        await ensure_logged_in(page, context, username, session_id=session_id)

        # Navigate to the Reels section. If the page ends up blank or showing
        # a grid, the user can click the Reels icon in Instagram's left sidebar.
        await safe_goto(page, f"{INSTAGRAM_URL}/reels/")
        await page.wait_for_timeout(3000)

        # If /reels/ didn't load a video (happens on fresh profiles), try clicking
        # the Reels nav icon directly so the player mode activates.
        try:
            has_video = await page.locator("video").count() > 0
            if not has_video:
                # Reels icon in the left nav
                reels_nav = page.locator('a[href="/reels/"]')
                if await reels_nav.count() > 0:
                    await reels_nav.first.click()
                    await page.wait_for_timeout(2000)
        except Exception:
            pass

        # Dismiss any popups that might be blocking the feed.
        await dismiss_popups(page)

        # From here the user is in control — let them scroll and train the feed.
        # Poll until all pages in the context are closed (user shut the window).
        while True:
            await page.wait_for_timeout(2000)
            try:
                if not context.pages:
                    break
            except Exception:
                break
        try:
            await context.close()
        except Exception:
            pass


def open_training_browser(username, session_id=None):
    """Blocking call: opens the trained browser for `username`, returns when closed."""
    asyncio.run(_open_training_browser_async(username, session_id=session_id))
